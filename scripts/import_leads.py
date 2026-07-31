"""Phase 2: import leads from an Excel file into the Google Sheet.

Usage:
    python -m scripts.import_leads <path_to_xlsx>

Reads `name` and `phone` columns (header names are matched case-insensitively;
if no recognizable header is found, the first two columns are assumed to be
name, phone). Phones are normalized to E.164 (+91 assumed for India). Each lead
is upserted with source='excel-import' and status='new'.
"""
import sys

from openpyxl import load_workbook

from app.sheets import upsert_lead
from app.utils import normalize_phone

NAME_KEYS = {"name", "lead name", "full name", "naam"}
PHONE_KEYS = {"phone", "mobile", "number", "phone number", "contact", "mobile number"}


def _detect_columns(header_row: list) -> tuple[int, int]:
    """Find (name_idx, phone_idx) from a header row, or default to 0,1."""
    name_idx = phone_idx = None
    for i, cell in enumerate(header_row):
        label = str(cell or "").strip().lower()
        if label in NAME_KEYS and name_idx is None:
            name_idx = i
        elif label in PHONE_KEYS and phone_idx is None:
            phone_idx = i
    if name_idx is None and phone_idx is None:
        return 0, 1  # no header detected -> assume first two columns
    return (name_idx or 0), (phone_idx if phone_idx is not None else 1)


def import_leads(path: str) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("No rows found in the file.")
        return 0

    # If the first row looks like a header (contains a known label), skip it.
    first = [str(c or "").strip().lower() for c in rows[0]]
    has_header = any(c in NAME_KEYS | PHONE_KEYS for c in first)
    name_idx, phone_idx = _detect_columns(rows[0]) if has_header else (0, 1)
    data_rows = rows[1:] if has_header else rows

    count = 0
    for row in data_rows:
        if not row or all(c is None for c in row):
            continue
        name = str(row[name_idx]).strip() if len(row) > name_idx and row[name_idx] else ""
        raw_phone = row[phone_idx] if len(row) > phone_idx else None
        phone = normalize_phone(raw_phone)
        if not phone or len(phone) < 8:
            print(f"  ! skipping row with bad phone: {row}")
            continue

        upsert_lead(
            {
                "phone": phone,
                "name": name,
                "source": "excel-import",
                "status": "new",
            }
        )
        print(f"  + {phone}  {name}")
        count += 1

    return count


def main() -> None:
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    path = sys.argv[1]
    print(f"Importing leads from: {path}")
    n = import_leads(path)
    print(f"\nDone. Imported/updated {n} lead(s) into the sheet.")


if __name__ == "__main__":
    main()
