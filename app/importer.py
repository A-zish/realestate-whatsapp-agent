"""Parsing uploaded lead lists (Excel or CSV).

Agencies export leads from Google Ads, Meta Lead Ads, a portal, or just keep
a spreadsheet — the column names are never consistent. This detects the name
and phone columns by header, falls back to position, and normalises every
phone to E.164 so the same lead from two sources doesn't become two rows.

Parsing is kept separate from persistence so the UI can show a preview and
let the user confirm before anything is written.
"""
import csv
import io
import logging

from openpyxl import load_workbook

from app.utils import normalize_phone

log = logging.getLogger(__name__)

NAME_HEADERS = {"name", "full name", "lead name", "customer name", "contact name", "naam", "first name"}
PHONE_HEADERS = {"phone", "phone number", "mobile", "mobile number", "number", "contact",
                 "contact number", "whatsapp", "whatsapp number", "phone_number", "mobile_no"}
EMAIL_HEADERS = {"email", "email address", "e-mail"}
CAMPAIGN_HEADERS = {"campaign", "campaign name", "campaign_name", "ad campaign", "utm campaign", "utm_campaign"}
GCLID_HEADERS = {"gclid", "google click id", "google_click_id", "click id"}
SOURCE_HEADERS = {"source", "lead source", "origin"}


def _norm_header(value) -> str:
    return str(value or "").strip().lower().replace("_", " ")


def _detect_columns(header_row: list) -> dict[str, int | None]:
    """Map logical fields to column indexes from a header row."""
    found: dict[str, int | None] = {
        "name": None, "phone": None, "campaign": None, "gclid": None, "source": None,
    }
    sets = {
        "name": NAME_HEADERS,
        "phone": PHONE_HEADERS,
        "campaign": CAMPAIGN_HEADERS,
        "gclid": GCLID_HEADERS,
        "source": SOURCE_HEADERS,
    }
    for i, cell in enumerate(header_row):
        label = _norm_header(cell)
        for key, headers in sets.items():
            if found[key] is None and label in headers:
                found[key] = i
    return found


def _rows_from_excel(content: bytes) -> list[list]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    return [list(r) for r in wb.active.iter_rows(values_only=True)]


def _rows_from_csv(content: bytes) -> list[list]:
    text = content.decode("utf-8-sig", errors="replace")
    return [row for row in csv.reader(io.StringIO(text))]


def parse_leads(filename: str, content: bytes) -> dict:
    """Parse an uploaded file into {leads, skipped, detected}.

    Never raises on a bad row — unusable rows land in `skipped` with a reason
    so the user can see exactly what was ignored and why.
    """
    name_l = filename.lower()
    if name_l.endswith((".xlsx", ".xlsm", ".xltx")):
        rows = _rows_from_excel(content)
    elif name_l.endswith(".csv"):
        rows = _rows_from_csv(content)
    else:
        raise ValueError("Unsupported file type — upload an .xlsx or .csv file.")

    rows = [r for r in rows if r and any(str(c or "").strip() for c in r)]
    if not rows:
        return {"leads": [], "skipped": [], "detected": "empty file"}

    cols = _detect_columns(rows[0])
    name_idx, phone_idx = cols["name"], cols["phone"]
    if name_idx is not None or phone_idx is not None:
        detected = (
            f"header row: name=column {(name_idx or 0) + 1}, "
            f"phone=column {(phone_idx or 1) + 1}"
        )
        extras = []
        if cols["campaign"] is not None:
            extras.append("campaign")
        if cols["gclid"] is not None:
            extras.append("gclid")
        if cols["source"] is not None:
            extras.append("source")
        if extras:
            detected += "; also " + ", ".join(extras)
        data_rows = rows[1:]
        name_idx = 0 if name_idx is None else name_idx
        phone_idx = 1 if phone_idx is None else phone_idx
    else:
        detected = "no header row found — using column 1 as name, column 2 as phone"
        name_idx, phone_idx = 0, 1
        data_rows = rows
        cols = {"campaign": None, "gclid": None, "source": None}

    leads, skipped, seen = [], [], set()
    for row in data_rows:
        raw_name = str(row[name_idx]).strip() if len(row) > name_idx and row[name_idx] else ""
        raw_phone = row[phone_idx] if len(row) > phone_idx else None
        phone = normalize_phone(raw_phone)

        if not phone or len(phone) < 10:
            skipped.append({"row": _preview(row), "reason": "no valid phone number"})
            continue
        if phone in seen:
            skipped.append({"row": _preview(row), "reason": f"duplicate of {phone}"})
            continue

        def _cell(idx: int | None) -> str:
            if idx is None or len(row) <= idx or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        seen.add(phone)
        leads.append({
            "name": raw_name,
            "phone": phone,
            "campaign": _cell(cols.get("campaign")),
            "gclid": _cell(cols.get("gclid")),
            "source": _cell(cols.get("source")),
        })

    return {"leads": leads, "skipped": skipped, "detected": detected}


def _preview(row: list) -> str:
    return ", ".join(str(c) for c in row[:3] if c is not None)[:60]
